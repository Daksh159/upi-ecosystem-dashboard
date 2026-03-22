"""
data_cleaner.py
---------------
Cleans and merges your 4 actual data files into 2 master CSVs:

OUTPUT 1: clean_upi_monthly.csv
  → 24 months of UPI volume + value (Apr 2023 – Mar 2025)
  → Source: NPCI 2023-24 + 2024-25 Excel files

OUTPUT 2: clean_payment_systems.csv
  → UPI vs NEFT vs IMPS vs RTGS comparison (Nov + Dec 2025)
  → Source: RBI PSI Excel files

Run:
  python scripts/data_cleaner.py
"""

import pandas as pd
import numpy as np
import re
import logging
from pathlib import Path
import openpyxl

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("logs/cleaning.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).resolve().parent.parent
RAW_DIR       = BASE_DIR / "data" / "raw" / "pdfs"   # your files live here
PROCESSED_DIR = BASE_DIR / "data" / "processed"
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
Path("logs").mkdir(exist_ok=True)


# ──────────────────────────────────────────────────────────────────────────────
# PART 1: Clean NPCI Monthly Files
# ──────────────────────────────────────────────────────────────────────────────

def clean_numeric_indian(value) -> float:
    """
    Indian number formats from NPCI files:
      '19,78,353.23'  → 1978353.23
      '18,301.51\t'   → 18301.51   (tab character present in 2024-25 file)
      '13440'         → 13440.0
      None            → NaN
    """
    if pd.isna(value):
        return np.nan
    
    val = str(value).strip()
    # Remove tabs, spaces, commas, ₹ symbol
    val = re.sub(r'[\t,₹\s]', '', val)
    
    try:
        return float(val)
    except ValueError:
        return np.nan


def parse_npci_month(month_str: str) -> pd.Timestamp:
    """
    NPCI month format: 'March-2024', 'April-2023', 'February-2025'
    Converts to proper datetime so SQL and Power BI can use it.
    """
    try:
        return pd.to_datetime(month_str.strip(), format="%B-%Y")
    except Exception:
        return pd.NaT


def clean_npci_file(filepath: Path) -> pd.DataFrame:
    """
    Reads one NPCI monthly Excel file and returns a clean DataFrame.
    
    Input columns:  Month | No. of Banks live on UPI | Volume (In Mn.) | Value (In Cr.)
    Output columns: report_month | banks_live | volume_mn | value_cr
    """
    logger.info(f"Reading NPCI file: {filepath.name}")
    
    df = pd.read_excel(filepath)
    
    # Rename columns to clean snake_case names
    df.columns = ["report_month", "banks_live", "volume_mn", "value_cr"]
    
    # Parse month string → datetime
    df["report_month"] = df["report_month"].apply(parse_npci_month)
    
    # Clean numeric columns
    df["banks_live"] = df["banks_live"].apply(clean_numeric_indian).astype("Int64")
    df["volume_mn"]  = df["volume_mn"].apply(clean_numeric_indian)
    df["value_cr"]   = df["value_cr"].apply(clean_numeric_indian)
    
    # Drop rows where month couldn't be parsed (usually empty rows)
    before = len(df)
    df = df.dropna(subset=["report_month"])
    after = len(df)
    
    if before != after:
        logger.info(f"  Dropped {before - after} unparseable rows")
    
    logger.info(f"  Rows extracted: {len(df)}")
    return df


def build_upi_monthly_master() -> pd.DataFrame:
    """
    Combines both NPCI files (2023-24 and 2024-25) into one 24-month master.
    Adds derived metrics useful for analysis.
    """
    npci_files = sorted(RAW_DIR.glob("*upi*Monthly*.xlsx"))
    
    if not npci_files:
        logger.error(
            "No NPCI files found. Make sure filenames contain 'upi' and "
            "'Monthly' and are in data/raw/pdfs/"
        )
        return pd.DataFrame()
    
    frames = [clean_npci_file(f) for f in npci_files]
    master = pd.concat(frames, ignore_index=True)
    
    # Sort chronologically oldest → newest
    master = master.sort_values("report_month").drop_duplicates(
        subset=["report_month"]
    ).reset_index(drop=True)
    
    # ── Derived metrics ──────────────────────────────────────────────────────
    
    # Month-over-month volume growth %
    # Formula: (current - previous) / previous * 100
    master["volume_mom_pct"] = master["volume_mn"].pct_change() * 100
    
    # Month-over-month value growth %
    master["value_mom_pct"] = master["value_cr"].pct_change() * 100
    
    # Average transaction value in Rupees
    # volume_mn is in millions, value_cr is in crores
    # 1 crore = 10 million, so avg value = (value_cr * 10^7) / (volume_mn * 10^6)
    # = value_cr * 10 / volume_mn
    master["avg_txn_value_rs"] = (master["value_cr"] * 10) / master["volume_mn"]
    
    # Year-over-year volume growth % (compare same month last year)
    master["volume_yoy_pct"] = master["volume_mn"].pct_change(periods=12) * 100
    
    # Add financial year label for grouping in Power BI
    # Indian FY runs April to March
    def get_fy(dt):
        if pd.isna(dt):
            return None
        if dt.month >= 4:
            return f"FY{dt.year}-{str(dt.year + 1)[2:]}"
        else:
            return f"FY{dt.year - 1}-{str(dt.year)[2:]}"
    
    master["financial_year"] = master["report_month"].apply(get_fy)
    
    # Round floats to 2 decimal places for cleanliness
    float_cols = ["volume_mn", "value_cr", "volume_mom_pct",
                  "value_mom_pct", "avg_txn_value_rs", "volume_yoy_pct"]
    master[float_cols] = master[float_cols].round(2)
    
    logger.info(f"\nUPI Monthly Master built:")
    logger.info(f"  Rows: {len(master)}")
    logger.info(f"  Date range: {master['report_month'].min().strftime('%b %Y')} "
                f"to {master['report_month'].max().strftime('%b %Y')}")
    logger.info(f"  Columns: {list(master.columns)}")
    
    return master


# ──────────────────────────────────────────────────────────────────────────────
# PART 2: Clean RBI Payment System Indicator Files
# ──────────────────────────────────────────────────────────────────────────────

# Payment systems we want to extract
# Key = exact label text as it appears in column B of the PSI file
PAYMENT_SYSTEM_LABELS = {
    "2.6 UPI @":                        "UPI",
    "2.3 IMPS":                         "IMPS",
    "2.5 NEFT":                         "NEFT",
    "1.1 Customer Transactions":        "RTGS",
    "4.1.1 PoS based $":               "Credit_Card",
    "4.2.1 PoS based $":               "Debit_Card",
}


def extract_month_label(filepath: Path) -> tuple:
    """
    Reads sheet name to get clean month label and report_month date.
    Sheet name is always 'Month YYYY' e.g. 'August 2024'
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet_name = wb.sheetnames[0]
    wb.close()

    try:
        dt = pd.to_datetime(sheet_name, format="%B %Y")
        report_month = dt.strftime("%Y-%m-01")
    except Exception:
        logger.warning(f"Could not parse sheet name: '{sheet_name}'")
        report_month = None

    return sheet_name, report_month


def scan_and_extract(filepath: Path) -> pd.DataFrame:
    """
    Scans every row in the PSI file looking for payment system labels.
    This approach is format-agnostic — works regardless of row shifts
    between old and new RBI layouts.

    Strategy:
    - Read col B (index 1) for the row label
    - Match against our PAYMENT_SYSTEM_LABELS dictionary
    - Extract col index 5 (current month volume) and col index 9 (value)

    Why this is better than hardcoded row numbers:
    RBI has changed row positions multiple times. Scanning by label
    means the script never breaks when RBI adds/removes rows.
    """
    sheet_name, report_month = extract_month_label(filepath)
    logger.info(f"Reading: {filepath.name} | Sheet: '{sheet_name}'")

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # Build a lookup: label_text -> (vol_lakh, val_cr)
    found = {}

    # Track Credit_Card and Debit_Card separately since both have
    # label "4.x.1 PoS based $" — we need to match in order
    credit_card_found = False

    for row in ws.iter_rows(max_row=60, values_only=True):
        # Col B = index 1 — the row label
        label = str(row[1]).strip() if row[1] is not None else ""

        for target_label, system_name in PAYMENT_SYSTEM_LABELS.items():
            if label == target_label:
                # Special handling: two rows have label "4.x.1 PoS based $"
                # First occurrence = Credit Card, Second = Debit Card
                if "PoS based" in label:
                    if not credit_card_found:
                        actual_system = "Credit_Card"
                        credit_card_found = True
                    else:
                        actual_system = "Debit_Card"
                else:
                    actual_system = system_name

                vol_raw = row[5] if len(row) > 5 else None
                val_raw = row[9] if len(row) > 9 else None

                vol = clean_numeric_indian(vol_raw)
                val = clean_numeric_indian(val_raw)

                # Volume in PSI files is in lakhs — convert to millions
                vol_mn = round(vol / 10, 2) if not np.isnan(vol) else np.nan
                val_cr = round(val, 2) if not np.isnan(val) else np.nan

                found[actual_system] = (vol_mn, val_cr)

                logger.info(
                    f"  {actual_system:15s} | "
                    f"Vol: {str(vol_mn):>12s} Mn | "
                    f"Val: {str(val_cr):>15s} Cr"
                )
                break  # found this label, move to next row

    wb.close()

    # Build records — include NaN rows for missing systems so
    # every file contributes the same columns
    records = []
    for system_name in ["UPI", "IMPS", "NEFT", "RTGS", "Credit_Card", "Debit_Card"]:
        vol_mn, val_cr = found.get(system_name, (np.nan, np.nan))
        records.append({
            "report_month":   report_month,
            "month_label":    sheet_name,
            "payment_system": system_name,
            "volume_mn":      vol_mn,
            "value_cr":       val_cr,
        })

    return pd.DataFrame(records)


def build_payment_systems_master() -> pd.DataFrame:
    """
    Processes all RBI PSI files into one master DataFrame.
    Handles both PSI* and PAYMENTSYSTEM* filename patterns.
    Adds UPI market share % per month.
    """
    psi_files = (
        sorted(RAW_DIR.glob("PSI*.XLSX")) +
        sorted(RAW_DIR.glob("PSI*.xlsx")) +
        sorted(RAW_DIR.glob("PAYMENTSYSTEM*.XLSX")) +
        sorted(RAW_DIR.glob("PAYMENTSYSTEM*.xlsx"))
    )

    if not psi_files:
        logger.error(
            "No RBI PSI files found. Place PSI*.XLSX or "
            "PAYMENTSYSTEM*.XLSX files in data/raw/pdfs/"
        )
        return pd.DataFrame()

    logger.info(f"\nFound {len(psi_files)} RBI PSI file(s)")

    frames = [scan_and_extract(f) for f in psi_files]
    master = pd.concat(frames, ignore_index=True)

    # Sort chronologically
    master["report_month"] = pd.to_datetime(master["report_month"])
    master = master.sort_values(
        ["report_month", "payment_system"]
    ).reset_index(drop=True)

    # Remove duplicates if any files overlap
    master = master.drop_duplicates(subset=["report_month", "payment_system"])

    # Add market share % per month
    monthly_totals = (
        master.groupby("report_month")[["volume_mn", "value_cr"]]
        .sum()
        .rename(columns={
            "volume_mn": "total_vol_mn",
            "value_cr":  "total_val_cr"
        })
        .reset_index()
    )

    master = master.merge(monthly_totals, on="report_month", how="left")

    master["share_of_volume_pct"] = (
        (master["volume_mn"] / master["total_vol_mn"]) * 100
    ).round(2)

    master["share_of_value_pct"] = (
        (master["value_cr"] / master["total_val_cr"]) * 100
    ).round(2)

    master = master.drop(columns=["total_vol_mn", "total_val_cr"])

    date_min = master["report_month"].min().strftime("%b %Y")
    date_max = master["report_month"].max().strftime("%b %Y")

    logger.info(f"\nPayment Systems Master:")
    logger.info(f"  Total rows     : {len(master)}")
    logger.info(f"  Months covered : {master['report_month'].nunique()}")
    logger.info(f"  Systems        : {sorted(master['payment_system'].unique())}")
    logger.info(f"  Date range     : {date_min} to {date_max}")

    return master
# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    logger.info("=" * 60)
    logger.info("Starting data cleaning pipeline")
    logger.info("=" * 60)
    
    # ── Part 1: NPCI Monthly ──────────────────────────────────────────────────
    upi_df = build_upi_monthly_master()
    
    if not upi_df.empty:
        out1 = PROCESSED_DIR / "clean_upi_monthly.csv"
        upi_df.to_csv(out1, index=False, encoding="utf-8-sig")
        logger.info(f"\nSaved: {out1}")
        print("\nPreview of clean_upi_monthly.csv:")
        print(upi_df.to_string())
    
    # ── Part 2: RBI Payment Systems ──────────────────────────────────────────
    psi_df = build_payment_systems_master()
    
    if not psi_df.empty:
        out2 = PROCESSED_DIR / "clean_payment_systems.csv"
        psi_df.to_csv(out2, index=False, encoding="utf-8-sig")
        logger.info(f"\nSaved: {out2}")
        print("\nPreview of clean_payment_systems.csv:")
        print(psi_df.to_string())
    
    logger.info("\nCleaning pipeline complete.")


if __name__ == "__main__":
    main()